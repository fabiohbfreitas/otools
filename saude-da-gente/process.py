# /// script
# dependencies = [
#   "openpyxl",
# ]
# ///

"""Organize SaudeDaGente xlsx exports into <dia>/<Especialidade>.csv import files.

Each xlsx file is a specialty; each sheet is a day (e.g. 24-08). Output lands in the
same folder as the input: <input_dir>/<sheet>/<Canonical>.csv with columns
Nome, Telefone, Etiquetas, Notas Internas.

Usage:
  uv run process.py <arquivo.xlsx>        # single file
  uv run process.py <pasta/>              # process *.xlsx in the folder (asks y/n per file)
  uv run process.py --daily <pasta/>      # also write combined daily files (2026-08-24.csv)
  uv run process.py --verbose <pasta/>    # extra per-sheet detail (columns, skips)
  uv run process.py --quant <file.xlsx>   # single-sheet variant: per-row Data+Especialidade,
                                          # output <YYYY-MM-DD>/<Canonical>.csv
  uv run process.py --validate <pasta/>   # compare outputs against inputs, full report
  uv run process.py --validate --daily <pasta/>  # also validate the combined daily files
  uv run process.py --duplicates <pasta/>     # flag shared phones across day-folder CSVs,
                                              # one <Especialidade>-duplicados.xlsx per specialty
  uv run process.py --duplicates <arquivo.csv>  # single CSV
  uv run process.py --selftest
"""
import csv
import re
import sys
import unicodedata
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

CRLF = "\r\n"
HEADERS = ["Nome", "Telefone", "Etiquetas", "Notas Internas"]
AUTOMACAO_TAG = "Automação"
SAUDE_TAG = "SaudeDaGente"
LOCAL_TAG = "Marajo"

ALIASES = [
    ("LABORATOR", "ExameLaboratorial"),
    ("CARDIOLOGISTA", "Cardiologia"),
    ("CARDIOLOGIA", "Cardiologia"),
    ("DERMATOLOGISTA", "Dermatologia"),
    ("DERMATOLOGIA", "Dermatologia"),
    ("US DE ABDOMEN TOTAL", "Ecografia"),
    ("US DE TIREOIDE", "Ecografia"),
    ("US MAMARIA", "Ecografia"),
    ("ECOGRAFIA", "Ecografia"),
    ("GINECOLOGISTA", "Ginecologia"),
    ("GINECOLOGIA", "Ginecologia"),
    ("MAMOGRAFIA", "Mamografia"),
    ("ORTOPEDISTA", "Ortopedia"),
    ("ORTOPEDIA", "Ortopedia"),
    ("PEDIATRA", "Pediatria"),
    ("PEDIATRIA", "Pediatria"),
    ("REUMATOLOGISTA", "Reumatologia"),
    ("REUMATOLOGIA", "Reumatologia"),
    ("ENDOCRINOLOGISTA", "Endocrinologia"),
    ("ENDOCRINOLOGIA", "Endocrinologia"),
    ("OTORRINOLARINGOLOGIA", "Otorrinolaringologia"),
    ("OTORRINO", "Otorrinolaringologia"),
    ("CLINICA MEDICA", "ClinicaMedica"),
    ("TOMOGRAFIA", "Tomografia"),
]

def normalize_token(s):
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^A-Z0-9]", "", s.upper())


NORMALIZED_ALIASES = [(normalize_token(t), c) for t, c in ALIASES]


def match_token(key, token):
    return token in key or key in token


def resolve_specialty(stem, workbook):
    norm = normalize_token(stem)
    for token, canonical in NORMALIZED_ALIASES:
        if match_token(norm, token):
            return canonical

    seen = set()
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if not header or header[0] is None:
            continue
        try:
            cols = resolve_columns(header)
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = normalize_token(str(row[cols["especialidade"]])) if row[cols["especialidade"]] is not None else ""
            if not key or key in seen:
                continue
            seen.add(key)
            for token, canonical in NORMALIZED_ALIASES:
                if match_token(key, token):
                    return canonical

    return norm.title()


def extract_phones(phone_str):
    if not phone_str:
        return []
    regex = re.compile(r"\(?\d{2}\)?\s?\d{4,5}-?\d{4}")
    result = []
    for num in regex.findall(phone_str):
        cleaned = re.sub(r"\D", "", num)
        if len(cleaned) == 11:
            result.append(f"({cleaned[0:2]}) {cleaned[2:7]}-{cleaned[7:]}")
        elif len(cleaned) == 10:
            result.append(f"({cleaned[0:2]}) {cleaned[2:6]}-{cleaned[6:]}")
        else:
            result.append(num.strip())
    return result


def pick_phone(phones):
    if not phones:
        return "", ""
    target = -1
    for i, p in enumerate(phones):
        digits = re.sub(r"\D", "", p)
        main = digits[2:]
        if len(main) == 9 and main.startswith("9"):
            target = i
            break
    if target == -1:
        target = 0
    chosen = phones[target]
    remaining = ", ".join(p for i, p in enumerate(phones) if i != target)
    return chosen, remaining


def split_trailing_phone(nome):
    """Split a trailing phone off a Nome cell.

    When the Telefone column is empty, the phone may be embedded at the end of
    the Nome cell (e.g. "Nome Do Paciente (61) 92345-1234"). Returns
    (name_part, phone_raw) with the trailing phone removed and normalized, or
    (nome, "") when there is no trailing phone. If stripping leaves no name,
    the name part is empty.
    """
    if not nome:
        return nome, ""
    regex = re.compile(r"\(?\d{2}\)?\s?\d{4,5}-?\d{4}")
    match = None
    for m in regex.finditer(nome):
        rest = nome[m.end():]
        if not rest or not rest.strip() or re.fullmatch(r"[\s\-()]+", rest):
            match = m
    if match is None:
        return nome, ""
    name_part = nome[:match.start()].strip().rstrip(" -()")
    phones = extract_phones(match.group(0))
    return name_part, phones[0] if phones else ""


def parse_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        month, day, year = m.groups()
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return ""


def build_row(nome, telefone, especialidade, date):
    phones = extract_phones(telefone)
    chosen, remaining = pick_phone(phones)
    etiquetas = f"{date}, {AUTOMACAO_TAG}, {especialidade}, {SAUDE_TAG}, {LOCAL_TAG}"
    notas = f"Outros telefones: {remaining}" if remaining else ""
    return {
        "Nome": nome,
        "Telefone": chosen,
        "Etiquetas": etiquetas,
        "Notas Internas": notas,
    }


def sanitize_folder(name):
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()


def parse_confirm(raw):
    """y/yes/s/sim/empty -> True; n/no/nao -> False; anything else -> None (re-ask)."""
    s = (raw or "").strip().lower()
    if s in ("", "y", "yes", "s", "sim"):
        return True
    if s in ("n", "no", "nao", "não"):
        return False
    return None


def confirm_file(name):
    while True:
        try:
            raw = input(f"Processar '{name}'? [Y/n] ")
        except EOFError:
            return False
        answer = parse_confirm(raw)
        if answer is not None:
            return answer


REQUIRED_COLUMNS = ["nome", "telefone", "data", "especialidade"]


def resolve_columns(header_values):
    """Map fuzzy header names to column indices (nome, telefone, data, especialidade)."""
    idx = {}
    for col_idx, raw in enumerate(header_values):
        key = normalize_token(str(raw)) if raw is not None else ""
        if not key:
            continue
        if "NOME" in key and "nome" not in idx:
            idx["nome"] = col_idx
        elif "TELEFONE" in key and "telefone" not in idx:
            idx["telefone"] = col_idx
        elif ("DATA" in key or "HORA" in key) and "data" not in idx:
            idx["data"] = col_idx
        elif "ESPECIALIDADE" in key and "especialidade" not in idx:
            idx["especialidade"] = col_idx
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise ValueError(
            f"Colunas não encontradas: {', '.join(missing)}. "
            f"Header lido: {header_values}"
        )
    return idx


def parse_sheet_date(sheet_name):
    m = re.match(r"^(\d{1,2})-(\d{1,2})$", sheet_name.strip())
    if not m:
        return ""
    day, month = int(m.group(1)), int(m.group(2))
    year = datetime.now().year
    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def build_sheet_rows(ws, cols, sheet_date, canonical):
    """Build output rows from one worksheet.

    Returns (rows, skipped, skipped_details) where rows is a list of
    (output_dict, spreadsheet_row_number, date), skipped is the number of
    rows dropped because they had no usable date, and skipped_details is a
    list of (spreadsheet_row_number, raw_data_value) for those dropped rows.
    """
    rows = []
    skipped = 0
    skipped_details = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        nome = row[cols["nome"]]
        if nome is None or not str(nome).strip():
            continue
        nome = str(nome).strip()
        date = parse_date(row[cols["data"]])
        if not date and sheet_date:
            date = sheet_date
        if not date:
            skipped += 1
            skipped_details.append((2 + i, row[cols["data"]]))
            continue
        telefone = str(row[cols["telefone"]]) if row[cols["telefone"]] is not None else ""
        if not telefone.strip():
            nome, telefone = split_trailing_phone(nome)
        out = build_row(nome, telefone, canonical, date)
        rows.append((out, 2 + i, date))
    return rows, skipped, skipped_details


def process_file(file_path, output_base, daily_rows=None, verbose=False):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    canonical = resolve_specialty(file_path.stem, workbook)
    created = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if not header or header[0] is None:
            if verbose:
                print(f"  {sheet}: planilha vazia, ignorada")
            continue

        try:
            cols = resolve_columns(header)
        except ValueError:
            if verbose:
                print(f"  {sheet}: cabeçalho não resolvido, ignorada")
            continue
        sheet_date = parse_sheet_date(sheet)
        sheet_rows, skipped, skipped_details = build_sheet_rows(ws, cols, sheet_date, canonical)
        if not sheet_rows and skipped == 0:
            continue
        if daily_rows is not None:
            for out, _, date in sheet_rows:
                daily_rows.setdefault(date, []).append(out)

        folder = output_base / sanitize_folder(sheet)
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / f"{canonical}.csv"
        existed = path.exists()
        if existed:
            path.unlink()
        write_csv_rows([out for out, _, _ in sheet_rows], path)
        created.append((len(sheet_rows), path, skipped, existed))
        if verbose:
            cols_str = ", ".join(f"{k}={v}" for k, v in sorted(cols.items()))
            print(f"  {sheet}: colunas {cols_str} | {len(sheet_rows)} linhas | {skipped} sem data")
            if skipped_details:
                cap = 10
                shown = skipped_details[:cap]
                for row_num, raw in shown:
                    print(f"    sem data: linha {row_num} ({raw})")
                rest = len(skipped_details) - len(shown)
                if rest > 0:
                    print(f"    ... mais {rest} linha(s) sem data")
    return canonical, created


def write_csv_rows(rows, path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=HEADERS, delimiter=",",
            quoting=csv.QUOTE_ALL, lineterminator=CRLF,
        )
        writer.writeheader()
        writer.writerows(rows)


def resolve_row_specialty(raw):
    """Canonical specialty for one Especialidade cell (quant variant)."""
    key = normalize_token(str(raw)) if raw is not None else ""
    if not key:
        return ""
    for token, canonical in NORMALIZED_ALIASES:
        if match_token(key, token):
            return canonical
    return key.title()


def process_file_quant(file_path, output_base, daily_rows=None, verbose=False):
    """Process the single-sheet per-row variant (Data + Especialidade on every row).

    Output goes to <output_base>/<YYYY-MM-DD>/<Canonical>.csv. Returns a list of
    (row_count, output_path, skipped_without_date, existed_before).
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    created = []
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if not header or header[0] is None:
            if verbose:
                print(f"  {sheet}: planilha vazia, ignorada")
            continue
        try:
            cols = resolve_columns(header)
        except ValueError:
            if verbose:
                print(f"  {sheet}: cabeçalho não resolvido, ignorada")
            continue

        groups = {}
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome = row[cols["nome"]]
            if nome is None or not str(nome).strip():
                continue
            nome = str(nome).strip()
            date = parse_date(row[cols["data"]])
            if not date:
                skipped += 1
                continue
            telefone = str(row[cols["telefone"]]) if row[cols["telefone"]] is not None else ""
            if not telefone.strip():
                nome, telefone = split_trailing_phone(nome)
            canonical = resolve_row_specialty(row[cols["especialidade"]])
            out = build_row(nome, telefone, canonical, date)
            groups.setdefault((date, canonical), []).append(out)
            if daily_rows is not None:
                daily_rows.setdefault(date, []).append(out)

        if not groups and skipped == 0:
            continue
        for (date, canonical), rows in sorted(groups.items()):
            folder = output_base / date
            folder.mkdir(parents=True, exist_ok=True)
            path = folder / f"{canonical}.csv"
            existed = path.exists()
            if existed:
                path.unlink()
            write_csv_rows(rows, path)
            created.append((len(rows), path, skipped, existed))
            if verbose:
                print(f"  {sheet}/{date}/{canonical}: {len(rows)} linhas | {skipped} sem data")
    return created


FIELDS = ["Nome", "Telefone", "Etiquetas", "Notas Internas"]


def etiquetas_date(row):
    et = row.get("Etiquetas", "")
    return et.split(",")[0].strip() if et else ""


def read_csv_rows(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [(row, 2 + i) for i, row in enumerate(reader)]


def canonical_from_stem(stem):
    norm = normalize_token(stem)
    for token, canonical in NORMALIZED_ALIASES:
        if match_token(norm, token):
            return canonical
    return norm.title()


def collect_row_phones(row):
    """All normalized phones of one CSV row, deduplicated, order preserved."""
    phones = []
    for field in ("Telefone", "Notas Internas", "Nome"):
        for p in extract_phones(row.get(field, "")):
            if p not in phones:
                phones.append(p)
    return phones


def name_key(nome):
    return " ".join(nome.split()).casefold()


def find_duplicates(rows):
    """Group occurrences by phone; keep groups sharing 2+ distinct names.

    rows: list of (row_dict, line_number). Returns [(phone, occurrences)]
    sorted by phone, where occurrences is a list of (nome, phones, date).
    """
    by_phone = {}
    for row, _ in rows:
        phones = collect_row_phones(row)
        nome = row.get("Nome", "")
        for p in phones:
            by_phone.setdefault(p, []).append((nome, phones, etiquetas_date(row)))
    flagged = [
        (phone, occs)
        for phone, occs in by_phone.items()
        if len({name_key(n) for n, _, _ in occs}) >= 2
    ]
    flagged.sort(key=lambda item: item[0])
    return flagged


def write_spreadsheet(specialty_groups, out_path):
    """Write one workbook with all specialties' flagged groups on a single sheet.

    specialty_groups: list of (canonical, flagged) where flagged is [(phone, occs)].
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicados"
    ws.append(["Nome", "Telefones", "Especialidade", "Data"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    first = True
    for canonical, flagged in specialty_groups:
        for _, occs in flagged:
            if not first:
                ws.append(["", "", "", ""])
            first = False
            for nome, phones, date in occs:
                ws.append([nome, ";".join(phones), canonical, date])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"
    try:
        wb.save(out_path)
    except PermissionError:
        print(f"Erro: feche '{out_path.name}' no Excel e tente novamente.")
        sys.exit(1)


DUPLICADOS_FILENAME = "duplicados.xlsx"


def collect_groups(path):
    """Map canonical specialty -> rows, from a folder of day folders or one CSV."""
    if path.is_file():
        return {canonical_from_stem(path.stem): read_csv_rows(path)}, path.parent
    day_dirs = [
        d for d in sorted(path.iterdir())
        if d.is_dir() and (re.match(r"^\d{1,2}-\d{1,2}$", d.name) or re.match(r"^\d{4}-\d{2}-\d{2}$", d.name))
    ]
    if not day_dirs:
        return {}, None
    groups = {}
    for day in day_dirs:
        for csv_path in sorted(day.glob("*.csv")):
            canonical = canonical_from_stem(csv_path.stem)
            groups.setdefault(canonical, []).extend(read_csv_rows(csv_path))
    return groups, path


def run_duplicates(path):
    """Detect patients sharing phone numbers in processed CSV outputs.

    path points at the output base containing day folders (or a single CSV).
    Writes a single duplicados.xlsx (all specialties on one sheet) next to the
    day folders. Returns exit code 0 on success, 1 on error.
    """
    if not path.exists():
        print(f"Caminho não encontrado: {path}")
        return 1

    groups, out_base = collect_groups(path)
    if not groups:
        print(f"Nenhuma pasta de dia (DD-MM) com CSVs encontrada em {path}")
        return 1

    flagged_any = []
    for canonical in sorted(groups):
        rows = groups[canonical]
        flagged = find_duplicates(rows)
        if not flagged:
            print(f"{canonical}: nenhum telefone duplicado ({len(rows)} linhas).")
            continue
        total_occs = sum(len(occs) for _, occs in flagged)
        print(f"{canonical}: {len(flagged)} número(s) duplicado(s), "
              f"{total_occs} linha(s) afetada(s)")
        for phone, occs in flagged:
            names = sorted({n for n, _, _ in occs})
            print(f"  {phone}: {len(occs)} linha(s), {len(names)} paciente(s): {', '.join(names)}")
        flagged_any.append((canonical, flagged))

    if not flagged_any:
        print("Nenhuma planilha gerada.")
        return 0
    out_path = out_base / DUPLICADOS_FILENAME
    write_spreadsheet(flagged_any, out_path)
    print(f"Planilha salva em: {out_path}")
    return 0


def compare_sheet(expected, actual):
    """Diff expected rows against actual CSV rows.

    expected: list of (output_dict, spreadsheet_row_number, date).
    actual:   list of (csv_dict, csv_line_number).

    Returns (missing, extra, diffs):
      missing = [(spreadsheet_row_number, output_dict)]
      extra   = [(csv_line_number, csv_dict)]
      diffs   = [(spreadsheet_row_number, field, expected_value, actual_value)]
    """
    buckets = defaultdict(deque)
    for idx, (arow, _) in enumerate(actual):
        buckets[(arow.get("Nome", ""), etiquetas_date(arow))].append(idx)

    matched = set()
    missing = []
    diffs = []
    for out, row_num, date in expected:
        q = buckets.get((out["Nome"], date))
        if q:
            idx = q.popleft()
            matched.add(idx)
            arow = actual[idx][0]
            for field in FIELDS:
                if out[field] != arow.get(field, ""):
                    diffs.append((row_num, field, out[field], arow.get(field, "")))
        else:
            missing.append((row_num, out))
    extra = [(2 + idx, arow) for idx, (arow, _) in enumerate(actual) if idx not in matched]
    return missing, extra, diffs


def _report_sheet(lines, totals, sheet_label, expected, csv_path):
    if not csv_path.exists():
        totals["no_file"] += 1
        totals["missing"] += len(expected)
        lines.append(f"  {sheet_label}: arquivo não encontrado: {csv_path.name}")
        lines.append(f"    FALTANDO {len(expected)} linha(s)")
        return
    actual = read_csv_rows(csv_path)
    missing, extra, diffs = compare_sheet(expected, actual)
    diff_rows = {rn for rn, _, _, _ in diffs}
    ok = len(expected) - len(missing) - len(diff_rows)
    totals["ok"] += ok
    totals["diff"] += len(diff_rows)
    totals["missing"] += len(missing)
    totals["extra"] += len(extra)
    lines.append(
        f"  {sheet_label}: OK {ok} | DIFERENTE {len(diff_rows)} | "
        f"FALTANDO {len(missing)} | EXTRA {len(extra)}"
    )
    for row_num, field, exp, act in diffs:
        lines.append(f"    DIFF {field} linha {row_num}: esperado '{exp}', obtido '{act}'")
    for row_num, out in missing:
        lines.append(
            f"    FALTANDO linha {row_num}: {out['Nome']} ({etiquetas_date(out)}) - "
            f"tel {out['Telefone']}"
        )
    for csv_line, arow in extra:
        lines.append(
            f"    EXTRA linha {csv_line}: {arow.get('Nome', '')} "
            f"({etiquetas_date(arow)}) - tel {arow.get('Telefone', '')}"
        )


def validate(path, daily):
    if path.is_dir():
        files = sorted(path.glob("*.xlsx"))
        output_base = path
    else:
        files = [path]
        output_base = path.parent
    if not files:
        print("Nenhum arquivo xlsx encontrado.")
        return 1

    lines = []
    totals = {"ok": 0, "diff": 0, "missing": 0, "extra": 0, "skipped": 0, "no_file": 0}
    daily_expected = {}

    for f in files:
        workbook = openpyxl.load_workbook(f, data_only=True)
        canonical = resolve_specialty(f.stem, workbook)
        lines.append(f"{f.name}  ->  {canonical}")
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
            if not header or header[0] is None:
                continue
            try:
                cols = resolve_columns(header)
            except ValueError:
                continue
            sheet_date = parse_sheet_date(sheet)
            sheet_rows, skipped, _ = build_sheet_rows(ws, cols, sheet_date, canonical)
            if not sheet_rows and skipped == 0:
                continue
            totals["skipped"] += skipped
            if daily:
                for out, row_num, date in sheet_rows:
                    daily_expected.setdefault(date, []).append((out, row_num, date))
            _report_sheet(
                lines, totals, sheet,
                sheet_rows,
                output_base / sanitize_folder(sheet) / f"{canonical}.csv",
            )

    if daily:
        lines.append("Diário:")
        for date in sorted(daily_expected):
            _report_sheet(
                lines, totals, f"{date}.csv",
                daily_expected[date],
                output_base / f"{date}.csv",
            )

    lines.append(
        f"\nTotal: OK {totals['ok']} | DIFERENTE {totals['diff']} | "
        f"FALTANDO {totals['missing']} | EXTRA {totals['extra']} | "
        f"SEM DATA {totals.get('skipped', 0)}"
    )
    if totals["no_file"]:
        lines.append(f"Aviso: {totals['no_file']} arquivo(s) de saída não encontrado(s).")

    report_path = output_base / "validation-report.md"
    with open(report_path, "w", encoding="utf-8") as fp:
        fp.write("# Relatório de validação\n\n")
        fp.write("\n".join(lines))
        fp.write("\n")
    print("\n".join(lines))
    print(f"\nRelatório salvo em: {report_path}")
    return 1 if (totals["diff"] or totals["missing"] or totals["extra"]) else 0


def selftest():
    assert extract_phones("(61) 99922-0084") == ["(61) 99922-0084"]
    assert extract_phones("(61) 98490-4799 (61) 98146-5163") == \
        ["(61) 98490-4799", "(61) 98146-5163"]
    assert extract_phones("(61) 3331-5174, (61) 99439-7065") == \
        ["(61) 3331-5174", "(61) 99439-7065"]
    assert extract_phones("61 99645-0163") == ["(61) 99645-0163"]
    chosen, remaining = pick_phone(["(61) 3331-5174", "(61) 99439-7065"])
    assert chosen == "(61) 99439-7065" and remaining == "(61) 3331-5174"
    chosen, remaining = pick_phone(["(61) 3331-5174"])
    assert chosen == "(61) 3331-5174" and remaining == ""
    assert parse_date("8/24/2026") == "2026-08-24"
    assert parse_date("2026-08-24 00:00:00") == "2026-08-24"
    assert parse_date(datetime(2026, 8, 24)) == "2026-08-24"
    assert parse_sheet_date("24-08") == "2026-08-24"
    assert parse_sheet_date("25-08") == "2026-08-25"
    assert parse_sheet_date("abc") == ""
    assert resolve_columns(["QUANT.", "Nome", "Telefone", "Data", "Hora", "Especialidade", "Local"]) == \
        {"nome": 1, "telefone": 2, "data": 3, "especialidade": 5}
    assert resolve_columns(["#", "NOME", "TELEFONE(S)", "DATA/HORA", "ESPECIALIDADE"]) == \
        {"nome": 1, "telefone": 2, "data": 3, "especialidade": 4}
    assert build_row("FULANO", "(61) 99988-7766, (61) 3622-3421",
                     "Cardiologia", "2026-08-24") == {
        "Nome": "FULANO",
        "Telefone": "(61) 99988-7766",
        "Etiquetas": "2026-08-24, Automação, Cardiologia, SaudeDaGente, Marajo",
        "Notas Internas": "Outros telefones: (61) 3622-3421",
    }
    assert etiquetas_date({"Etiquetas": "2026-08-24, Automação, Cardiologia, SaudeDaGente, Marajo"}) == "2026-08-24"
    expected = [
        (build_row("A", "(61) 99999-0001", "Cardiologia", "2026-08-24"), 3, "2026-08-24"),
        (build_row("B", "(61) 99999-0002", "Cardiologia", "2026-08-24"), 4, "2026-08-24"),
    ]
    actual = [
        (build_row("A", "(61) 99999-0001", "Cardiologia", "2026-08-24"), 2),
        (build_row("C", "(61) 99999-0003", "Cardiologia", "2026-08-24"), 3),
    ]
    missing, extra, diffs = compare_sheet(expected, actual)
    assert len(missing) == 1 and missing[0][1]["Nome"] == "B"
    assert len(extra) == 1 and extra[0][1]["Nome"] == "C"
    assert diffs == []
    missing, extra, diffs = compare_sheet(expected[:1], [
        (build_row("A", "(61) 99999-9999", "Cardiologia", "2026-08-24"), 2),
    ])
    assert not missing and not extra
    assert any(f == "Telefone" for _, f, _, _ in diffs)
    assert split_trailing_phone("Nome Do Paciente (61) 92345-1234") == \
        ("Nome Do Paciente", "(61) 92345-1234")
    assert split_trailing_phone("Nome Do Paciente 61 92345-1234") == \
        ("Nome Do Paciente", "(61) 92345-1234")
    assert split_trailing_phone("Nome Do Paciente (61) 3345-1234") == \
        ("Nome Do Paciente", "(61) 3345-1234")
    assert split_trailing_phone("(61) 92345-1234") == ("", "(61) 92345-1234")
    assert split_trailing_phone("Nome Do Paciente") == ("Nome Do Paciente", "")
    assert split_trailing_phone("Maria (61) 92345-1234 Silva") == \
        ("Maria (61) 92345-1234 Silva", "")
    assert split_trailing_phone("Nome (61) 92345-1234 (61) 3345-1234") == \
        ("Nome (61) 92345-1234", "(61) 3345-1234")
    assert canonical_from_stem("Cardiologia") == "Cardiologia"
    assert canonical_from_stem("Laboratório") == "ExameLaboratorial"
    assert canonical_from_stem("Endocrinologista") == "Endocrinologia"
    assert canonical_from_stem("Tomografia") == "Tomografia"
    assert collect_row_phones({
        "Telefone": "(61) 99988-7766",
        "Notas Internas": "Outros telefones: (61) 3622-3421",
        "Nome": "FULANO",
    }) == ["(61) 99988-7766", "(61) 3622-3421"]
    assert collect_row_phones({
        "Telefone": "(61) 99988-7766",
        "Notas Internas": "",
        "Nome": "X",
    }) == ["(61) 99988-7766"]
    assert collect_row_phones({"Telefone": "", "Notas Internas": "", "Nome": "Sem telefone"}) == []
    assert collect_row_phones({"Telefone": "61998708050", "Notas Internas": "", "Nome": ""}) == \
        ["(61) 99870-8050"]

    dup_rows = [
        ({"Nome": "ANA SILVA", "Telefone": "(61) 99999-0001", "Etiquetas": ""}, 2),
        ({"Nome": "ana  silva", "Telefone": "(61) 99999-0001", "Etiquetas": ""}, 3),
        ({"Nome": "BRUNO COSTA", "Telefone": "(61) 99999-0002",
          "Etiquetas": "2026-08-24, Automação, Cardiologia, SaudeDaGente, Marajo"}, 4),
        ({"Nome": "CARLOS DIAS", "Telefone": "(61) 99999-0002",
          "Etiquetas": "2026-08-25, Automação, Cardiologia, SaudeDaGente, Marajo"}, 5),
        ({"Nome": "DUDA SOUZA", "Telefone": "(61) 99999-0003", "Etiquetas": ""}, 6),
        ({"Nome": "DUDA SOUZA", "Telefone": "(61) 99999-0003", "Etiquetas": ""}, 7),
    ]
    flagged = find_duplicates(dup_rows)
    assert [p for p, _ in flagged] == ["(61) 99999-0002"]
    g1 = dict(flagged)["(61) 99999-0002"]
    assert len(g1) == 2 and g1[0][0] == "BRUNO COSTA" and g1[1][0] == "CARLOS DIAS"
    assert g1[0][2] == "2026-08-24" and g1[1][2] == "2026-08-25"

    assert resolve_row_specialty("GINECOLOGISTA") == "Ginecologia"
    assert resolve_row_specialty("Cardiologia") == "Cardiologia"
    assert resolve_row_specialty("Exames laboratoriais") == "ExameLaboratorial"
    assert resolve_row_specialty("Fono") == "Fono"

    import tempfile
    with tempfile.TemporaryDirectory() as td:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["QUANT.", "Nome", "Telefone", "Data", "Hora", "Especialidade", "Local"])
        ws.append([1, " MARIA APARECIDA", "(61) 98194-5975",
                   datetime(2026, 8, 26), datetime(2026, 8, 26, 7, 0),
                   "GINECOLOGISTA", "sáude da gente - marajo "])
        ws.append([2, "SEM DATA", "(61) 98194-5975", "", "", "GINECOLOGISTA", "x"])
        sample = Path(td) / "26.08 GINECOLOGISTA.xlsx"
        wb.save(sample)
        quant_daily = {}
        created = process_file_quant(sample, Path(td), quant_daily)
        assert len(created) == 1
        count, out_path, skipped, existed = created[0]
        assert count == 1 and skipped == 1 and existed is False
        assert out_path == Path(td) / "2026-08-26" / "Ginecologia.csv"
        rows = read_csv_rows(out_path)
        assert len(rows) == 1
        row = rows[0][0]
        assert row["Nome"] == "MARIA APARECIDA"
        assert row["Telefone"] == "(61) 98194-5975"
        assert row["Etiquetas"] == "2026-08-26, Automação, Ginecologia, SaudeDaGente, Marajo"
        assert list(quant_daily) == ["2026-08-26"]
        again = process_file_quant(sample, Path(td))
        assert again[0][0] == 1 and again[0][3] is True

    assert parse_confirm("") is True
    assert parse_confirm(" y ") is True
    assert parse_confirm("S") is True
    assert parse_confirm("sim") is True
    assert parse_confirm("n") is False
    assert parse_confirm("NÃO") is False
    assert parse_confirm("x") is None

    with tempfile.TemporaryDirectory() as td:
        dup_path = Path(td) / DUPLICADOS_FILENAME
        write_spreadsheet([
            ("Cardiologia", [("(61) 99999-0002", [
                ("BRUNO COSTA", ["(61) 99999-0002"], "2026-08-24"),
                ("CARLOS DIAS", ["(61) 99999-0002"], "2026-08-25"),
            ])]),
            ("Pediatria", [("(61) 99999-0009", [
                ("DUDA SOUZA", ["(61) 99999-0009"], "2026-08-26"),
            ])]),
        ], dup_path)
        wb = openpyxl.load_workbook(dup_path)
        ws = wb.active
        assert ws.title == "Duplicados" and len(wb.sheetnames) == 1
        assert [c.value for c in ws[1]] == ["Nome", "Telefones", "Especialidade", "Data"]
        assert ws["A2"].value == "BRUNO COSTA" and ws["B2"].value == "(61) 99999-0002"
        assert ws["C2"].value == "Cardiologia" and ws["D3"].value == "2026-08-25"
        assert ws["A4"].value is None  # blank separator between groups
        assert ws["A5"].value == "DUDA SOUZA" and ws["C5"].value == "Pediatria"
    print("selftest ok")


def main():
    if "--selftest" in sys.argv:
        selftest()
        return

    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    daily = "--daily" in sys.argv
    quant = "--quant" in sys.argv
    if len(args) < 1:
        print("Uso: uv run process.py [--validate] [--daily] [--verbose] "
              "[--duplicates] [--quant] <arquivo|pasta>")
        sys.exit(1)

    path = Path(args[0])
    verbose = "--verbose" in sys.argv
    if "--duplicates" in sys.argv:
        for flag in ("--daily", "--validate", "--verbose", "--quant"):
            if flag in sys.argv:
                print(f"Erro: --duplicates não pode ser combinado com {flag}.")
                sys.exit(1)
        sys.exit(run_duplicates(path))

    if "--validate" in sys.argv:
        if quant:
            print("Erro: --quant não pode ser combinado com --validate.")
            sys.exit(1)
        sys.exit(validate(path, daily))

    if path.is_dir():
        files = sorted(path.glob("*.xlsx"))
        output_base = path
    else:
        files = [path]
        output_base = path.parent

    declined = []
    if path.is_dir():
        confirmed = []
        for f in files:
            if confirm_file(f.name):
                confirmed.append(f)
            else:
                declined.append(f)
        files = confirmed

    if not files:
        if declined:
            print("Nenhum arquivo selecionado.")
            sys.exit(0)
        print("Nenhum arquivo xlsx encontrado.")
        sys.exit(1)

    daily_rows = {} if daily else None

    if quant:
        grand_total = 0
        grand_skipped = 0
        for f in files:
            created = process_file_quant(f, output_base, daily_rows, verbose)
            total = sum(n for n, _, _, _ in created)
            skipped = sum(s for _, _, s, _ in created)
            print(f"{total:5d}  {f.name}")
            for count, p, skip, existed in created:
                line = f"      {count:5d}  {p}"
                if skip:
                    line += f"  ({skip} sem data)"
                line += "  [sobrescrito]" if existed else "  [novo]"
                print(line)
            grand_total += total
            grand_skipped += skipped
    else:
        grand_total = 0
        grand_skipped = 0
        for f in files:
            canonical, created = process_file(f, output_base, daily_rows, verbose)
            total = sum(n for n, _, _, _ in created)
            skipped = sum(s for _, _, s, _ in created)
            print(f"{total:5d}  {f.name}  ->  {canonical}")
            for count, p, skip, existed in created:
                line = f"      {count:5d}  {p}"
                if skip:
                    line += f"  ({skip} sem data)"
                line += "  [sobrescrito]" if existed else "  [novo]"
                print(line)
            grand_total += total
            grand_skipped += skipped
    print(f"Total: {grand_total} linhas")
    if grand_skipped:
        print(f"Aviso: {grand_skipped} linha(s) ignorada(s) por não terem data.")

    if daily:
        for date, rows in sorted(daily_rows.items()):
            path = output_base / f"{date}.csv"
            existed = path.exists()
            write_csv_rows(rows, path)
            mark = "[sobrescrito]" if existed else "[novo]"
            print(f"Diário: {len(rows):5d}  {path}  {mark}")

    if declined:
        print("\nIgnorados pelo usuário:")
        for f in declined:
            print(f"  {f.name}")


if __name__ == "__main__":
    main()